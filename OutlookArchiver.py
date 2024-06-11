# Outlook Archiver

# 2023 - 2024 Niels Buring
# Versie 0.3

# Laatste update: 2024 06 11

# Changelog

# 0.3
# Disabled excel feature (uncomment to use)

# 0.2
# Added Excel feature

# 0.1
# Initial release


import win32com.client
import os
import re

import openpyxl
from openpyxl import load_workbook

from datetime import datetime, timedelta


import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import *


# filename= r"C:\Users\Paulx\OneDrive\Bureaublad\DocumentenNiels\Scripts\test.xlsx"
# workbook = load_workbook(filename)
# sheet = workbook.active

outlook = win32com.client.Dispatch("Outlook.Application")
messages = outlook.ActiveExplorer().Selection
count = outlook.ActiveExplorer().Selection.Count + 1

print('initializing')
print('\n')
print("selected emails: " + str(count))

win = Tk()
win.title('EmailArchiver 0.3')


#  win.geometry("150x180")  #size with outlook
win.geometry("150x120")  #size with outlook


def quit():
    win.quit()
    win.destroy()


# Define empty variables
var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()

# Define a Checkbox
l1 = Label(win, text="What to archive?", justify="right")
l1.grid(sticky = W, pady=5)

#l1.pack()
t1 = Checkbutton(win, text="Email", variable=var1, onvalue=1, offvalue=0, justify="left")
#t1.pack()
t1.grid(sticky = W)
t2 = Checkbutton(win, text="Attachments", variable=var2, onvalue=1, offvalue=0, justify="right")
#t2.pack()
t2.grid(sticky = W)


s = ttk.Separator(win, orient=HORIZONTAL)
s.grid(sticky = W, ipadx=150)



# l2 = Label(win, text="Excel?")
# #l2.pack(padx=50)
# l2.grid(sticky = W, pady=5)
# t3 = Checkbutton(win, text="Excel", variable=var3, onvalue=1, offvalue=0)
# #t3.pack()
# t3.grid(sticky = W)




# TEST

# t4 = Checkbutton(win, text="Custom name?", variable=var4, onvalue=1, offvalue=0)
# t4.pack()
# e1 = Entry(win)
# e1.pack()

# entry = e1.get()
# print("INVOER:")
# print(entry)


b1 = Button(win, text="Continue", command=quit)
#b1.pack()
b1.grid(sticky = W, pady=5)


win.mainloop()


if var1.get() == 1 or var2.get() == 1:

    folder = filedialog.askdirectory(title="Select Save Path")
    path = folder

for x in range(1,count):
    
    print("Bericht:"+ str(x))
    print(messages(x))
    message = messages(x)
    
    #change category
    if var1.get() == 1 or var2.get() == 1:

        message.Categories= 'Opgeslagen'
        message.Save()

    #naam
    name = str(message.subject)
    name = re.sub('[^A-Za-z0-9]+', ' ', name)

    #datum
    delta = 7

    time = message.ReceivedTime
    timeDue = time + timedelta(days=delta)

    date = time.strftime("%Y%m%d")
    dateDue= timeDue.strftime("%d-%m-%Y")

    if var1.get() == 1 or var2.get() == 1:

        pathNew = path+'/'+date+' - '+name
        exist = os.path.exists(pathNew)
        if exist == False:
            os.mkdir(pathNew)
        print(pathNew)



    #########################
    #       save email      #
    #########################
            
    if var1.get() == 1:
        
        nameTest = name+'.msg'
        isFile = os.path.isfile(pathNew+'//'+nameTest)
        if isFile == True:
            name = name+"_1"+".msg"
        else:
            name = name+'.msg'
        
        message.SaveAs(pathNew+'//'+name)
        print(name)


    #########################
    #   save attachments    #
    #########################
    
    if var2.get() == 1:

        for attachment in message.Attachments:
            attName = str(attachment)
            attachment.SaveAsFile(os.path.join(pathNew, attName))

    #########################
    #          excel        #
    #########################

    #write excel

    if var3.get() == 1:
        
        rowE=sheet.max_row+1
        row="A"+str(rowE)
        rowDue="C"+str(rowE)

        sheet[row] = date
        sheet[rowDue] = dateDue
        workbook.save(filename=filename)



#input("Press enter to exit;")


print("Email:", var1.get())
print("Attachments:", var2.get())
print("Excel:", var3.get())



print("Done")

